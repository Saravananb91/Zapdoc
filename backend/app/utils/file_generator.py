
import openpyxl
from openpyxl.utils import get_column_letter
import tempfile
import os

def generate_excel_report(extracted_data: dict, pages: list = None) -> str:
    """
    Generates an Excel report from the extracted data.
    Returns the path to the temporary Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extraction Results"

    # 1. Header Info (Invoice Details)
    # We want a vertical key-value list for the header to fit more data cleanly
    
    # Define fields to show
    header_fields = [
        ("Invoice Number", extracted_data.get("invoice_no")),
        ("Date", extracted_data.get("date_of_issue")),
        ("Seller Name", extracted_data.get("seller_name")),
        ("Seller Address", extracted_data.get("seller_address")),
        ("Seller Tax ID", extracted_data.get("seller_tax_id")),
        ("Seller IBAN", extracted_data.get("seller_iban")),
        ("Client Name", extracted_data.get("client_name")),
        ("Client Address", extracted_data.get("client_address")),
        ("Client Tax ID", extracted_data.get("client_tax_id")),
        ("Client IBAN", extracted_data.get("client_iban")),
        ("Net Total", extracted_data.get("net_total")),
        ("VAT Total", extracted_data.get("vat_total")),
        ("Gross Total", extracted_data.get("gross_total")),
    ]

    for label, value in header_fields:
        ws.append([label, value])
    
    # Separator
    ws.append([])
    ws.append(["Line Items"])
    
    # 2. Line Items
    item_headers = [
        "Item No", "Description", "HSN", "Quantity", "Unit", "Unit Price", 
        "Net Amount", "Discount", "Tax Amount", "VAT Rate", "Gross Amount"
    ]
    ws.append(item_headers)
    
    items = extracted_data.get("items") or []
    if items:
        for item in items:
            ws.append([
                item.get("item_no"),
                item.get("description"),
                item.get("hsn_code"),
                item.get("qty"),
                item.get("unit"),
                item.get("unit_price") or item.get("rate"),
                item.get("net_amount"),
                item.get("discount"),
                item.get("tax_amount"),
                item.get("vat_rate"),
                item.get("total") or item.get("gross_amount"),
            ])
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap width

    # 3. Custom Fields (if any)
    custom_fields = extracted_data.get("custom_fields") or {}
    if custom_fields and isinstance(custom_fields, dict):
        ws.append([])
        ws.append(["Additional Extracted Data"])
        ws.append(["Field Name", "Value"])
        
        for key, value in custom_fields.items():
            ws.append([key.replace('_', ' ').title(), value])

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    
    return tmp.name
